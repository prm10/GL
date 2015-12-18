# -*- coding: utf-8 -*-
__author__ = 'prm14'
import xlrd
import openpyxl
import re

data = xlrd.open_workbook(u'二号炉产量和物理热范围.xlsx')
table1 = data.sheet_by_name('Sheet1')
book=openpyxl.Workbook()
sheet=book.get_sheet_by_name(name='Sheet')

pattern=re.compile('^(\d+)-(\d+)$')
value=0
for i in range(table1.nrows):
    for j in range(table1.ncols):
        value=table1.cell(i,j).value
        sheet.cell(row=i+1,column=j+1,value=value)
    match=pattern.match(value)
    if match:
        sheet.cell(row=i+1,column=table1.ncols+1,value=int(match.group(1)))
        sheet.cell(row=i+1,column=table1.ncols+2,value=int(match.group(2)))
    else:
        print(u'第%d行第%d列不符合规范'%(i+1,table1.ncols))
book.save('result.xlsx')
